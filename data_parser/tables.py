import openpyxl


def write_data_to_excel(product_data, file_name):
    # Создаем новую рабочую книгу
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Записываем заголовки (ключи словаря) в первую строку
    headers = list(product_data.keys())
    for col, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col, value=header)

    # Записываем значения (значения ключей словаря) во вторую строку
    values = list(product_data.values())
    for col, value in enumerate(values, start=1):
        if isinstance(value, list):
            # Если значение - это список, записываем его элементы в разные ячейки во второй строке
            for row_offset, item in enumerate(value, start=1):
                sheet.cell(row=2 + row_offset, column=col, value=item)
        else:
            sheet.cell(row=2, column=col, value=value)

    # Сохраняем книгу в файл
    workbook.save(file_name)
