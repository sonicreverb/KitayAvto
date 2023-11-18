import os.path
import database.postgres_connector as db
import ftplib
import openpyxl
from pycbrf import ExchangeRates

import logs
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.dirname(__file__))


# получаем текущий курс юаней
def get_cny_rate():
    try:
        rates = ExchangeRates()
        cny_to_rub_rate = rates['CNY'].value
        return float(cny_to_rub_rate)
    except Exception as _ex:
        logs.log_error(f"[GET CNY RATE] Error while trying to get current course {_ex}")
        raise SystemExit(-1)


# запись данных из БД в xlsx файл querry - запрос в БД, filename - название файла в который запись
def write_data_to_xlsx(querry, filename):
    # получаем соединение
    connection = db.get_connection_to_db()

    cny_rate = get_cny_rate()

    # если соединение установлено успешно
    if connection:
        with connection.cursor() as cursor:
            # SQL запрос
            cursor.execute(querry)
            rows = cursor.fetchall()

            # создание XLSX-файла
            workbook = openpyxl.Workbook()
            worksheet = workbook.active

            # запись заголовков столбцов
            columns = [i[0] for i in cursor.description]
            for column_index, column_name in enumerate(columns, start=1):
                column_letter = get_column_letter(column_index)
                worksheet[f"{column_letter}1"] = column_name
                worksheet["CI1"] = 'CNY_rate'

            # запись данных
            for row_index, row in enumerate(rows, start=2):
                for column_index, cell_value in enumerate(row):
                    column_letter = get_column_letter(column_index + 1)
                    try:
                        # для изображений изменяем хост (наш фтп)
                        if 7 <= column_index <= 21:
                            cell_value = "ftp://84.42.46.187" + cell_value[cell_value.rfind('/'):]
                        worksheet[f"{column_letter}{row_index}"] = str(cell_value)
                    except Exception as _ex:
                        print('[XLSX FILE] could\'t write cell value, string error.', _ex)
                        logs.log_warning(f'[XLSX FILE] could\'t write cell value, string error. {_ex}')
                    worksheet[f"CI{row_index}"] = str(cny_rate)

            # сохранение файла
            filename_path = os.path.join(BASE_DIR, 'database', 'output', filename)
            workbook.save(filename_path)

        connection.commit()
        # прикрываем соединение
        connection.close()
        print("[PostGreSQL INFO] Connection closed.")
        print("[XLSX FILE] XLSX file was updated successfully.")
        logs.log_info("[XLSX FILE] XLSX file was updated successfully.")
    else:
        print("[PostGreSQL INFO] Error, couldn't get connection...")


# загружает filename на фтп
def upload_file_to_ftp(filename):
    ftp_server = '84.42.46.187'
    ftp_username = 'ivvsavqz_3'
    ftp_password = 'O96QOw&y'
    remote_file_path = filename
    local_file_path = os.path.join(BASE_DIR, 'database', 'output', remote_file_path)

    # подключение к FTP серверу
    with ftplib.FTP(ftp_server, ftp_username, ftp_password) as ftp:
        print('[FTP INFO] Connecting to FTP server...')
        logs.log_info('[FTP INFO] Connecting to FTP server...')
        # открытие файла для чтения
        with open(local_file_path, 'rb') as file:
            # загрузка файла на сервер
            ftp.storbinary(f'STOR {remote_file_path}', file)
            print(f'[FTP INFO] {filename} was uploaded successfully!')
            logs.log_info(f'[FTP INFO] {filename} was uploaded successfully!')


TMP_IMGS_DIR = os.path.join(BASE_DIR, 'database', 'tmp_images')


# загружает массив изображений на фтп
def upload_imgs_to_ftp(imgs):
    ftp_server = '84.42.46.187'
    ftp_username = 'ivvsavqz_3'
    ftp_password = 'O96QOw&y'

    # подключение к FTP серверу
    with ftplib.FTP(ftp_server, ftp_username, ftp_password) as ftp:
        print('[FTP INFO] Connecting to FTP server...')
        logs.log_info('[FTP INFO] Connecting to FTP server...')
        # открытие файла для чтения
        for img in imgs:
            remote_file_path = img[img.rfind('/') + 1:]
            local_file_path = os.path.join(TMP_IMGS_DIR, remote_file_path)
            with open(local_file_path, 'rb') as file:
                # загрузка файла на сервер
                ftp.storbinary(f'STOR {remote_file_path}', file)
                # logs.log_info(f'{file} was uploaded successfully.')
            print(f'[FTP INFO] {remote_file_path} was uploaded successfully!')
            logs.log_info(f'[FTP INFO] {remote_file_path} was uploaded successfully!')


# удаляет всё содержимое в TMP_IMGS_DIR
def delete_tmp_imgs():
    for file in os.listdir(TMP_IMGS_DIR):
        try:
            file_path = os.path.join(TMP_IMGS_DIR, file)
            if os.path.isfile(file_path):
                os.unlink(file_path)
        except Exception as _ex:
            print(f'[DELETE TMP IMGS] Images was successfully deleted. An error occupied {_ex}')
            logs.log_warning(f'[DELETE TMP IMGS] Images was successfully deleted. An error occupied {_ex}')
    logs.log_info('[DELETE TMP IMGS] Images was successfully deleted.')
